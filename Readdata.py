import requests
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
# from openpyxl.styles import PatternFill, Border, Side, numbers
import sys
from bs4 import BeautifulSoup
import time
import datetime

class Timer(object):
    def __init__(self):
        self.start = datetime.datetime.now()

    def remains(self, done):
        if done != 0:
            now  = datetime.datetime.now()
            left = (1 - done) * (now - self.start) / done
            sec = int(left.total_seconds())
            if sec < 60:
                return "{} seconds   ".format(sec)
            else:
                return "{} minutes   ".format(int(sec / 60))

class ProgressBar(object):
    def __init__(self):
        self.timer = Timer()
        self.barLength = 20
        self.progress = 0.0

    def update(self, done):
        self.progress = float(done)
        block = int(round(self.barLength * self.progress))
        text = "\r[{0}] {1:.2f}% {2}".format( "="*block + " "*(self. barLength-block), self. progress*100, self.timer.remains(self.progress))
        sys.stdout.write(text)
        sys.stdout.flush()

class Pelicula(object):
    def __init__(self, movie_box):
            self.titulo = self.get_title(movie_box)
            self.user_note = self.get_user_note(movie_box)
            self.id = self.get_id(movie_box)
            self.url_FA = get_url_from_id(self.id)

            self.parsed_page = None
            self.nota_FA = 0
            self.votantes_FA = 0
            self.duracion = 0

    def get_title(self, film):
        return film.contents[1].contents[1].contents[3].contents[1].contents[0].contents[0]

    def get_user_note(self, film):
        return film.contents[3].contents[1].contents[1].contents[0]

    def get_id(self, film):
        return film.contents[1].contents[1].attrs['data-movie-id']

    def get_FA_url(self):
        return self.url_FA

    def valid(self):
        return es_valida(self.titulo)

    def get_parsed_page(self):
        resp = safe_get_url(self.get_FA_url())
        if resp.status_code == 404:
            self.exists = False
            return # Si el id no es correcto, dejo de construir la clase
        else:
            self.exists = True

        # Parseo la página
        self.parsed_page = BeautifulSoup(resp.text,'html.parser')

    def get_time_and_FA(self):
        if not self.parsed_page:
            self.get_parsed_page()
        # Comienzo a leer datos de la página
        l = self.parsed_page.find(id="movie-rat-avg")
        try:
            # guardo la nota de FA en una ficha normal
            self.nota_FA = float(l.attrs['content'])
        except:
            # caso en el que no hay nota de FA
            self.nota_FA = 0

        l = self.parsed_page.find(itemprop="ratingCount")
        try:
            # guardo la cantidad de votantes en una ficha normal
            self.votantes_FA = l.attrs['content']
            # Elimino el punto de los millares
            self.votantes_FA = self.votantes_FA.replace('.','')
            self.votantes_FA = int(self.votantes_FA)
        except:
            # caso en el que no hay suficientes votantes
            self.votantes_FA = 0

        l = self.parsed_page.find(id = "left-column")
        try:
            self.duracion = l.find(itemprop="duration").contents[0]
        except:
            # caso en el que no está escrita la duración
            self.duracion = "0"
        # quito el sufijo min.
        self.duracion = int(self.duracion.split(' ', 1)[0])

def es_valida(titulo):
    """
    Busca en el título que sea una película realmente
    """
    # Comprobamos que no tenga ninuno de los sufijos a evitar
    if titulo.find("(C)") > 0: # Excluyo los cortos
        return False
    if titulo.find("(Miniserie de TV)") > 0: # Excluyo series de televisión
        return False
    if titulo.find("(Serie de TV)") > 0:
        return False
    if titulo.find("(TV)") > 0:
        # Hay varios tipos de películas aquí.
        # Algunos son programas de televisón, otros estrenos directos a tele.
        # Hay también episosios concretos de series.
        return False
    if titulo.find("(Vídeo musical)") > 0:
        return False
    # No se ha encontrado sufijo, luego es una película
    return True

class IndexLine():
    # Esta clase debería ser capaz de dar cualquier orden a las películas
    def __init__(self, totales):
        self.m_totales = totales
        # Hay que inicializarlo para que no escriba en los encabezados
        self.m_current = 2

    def get_current_line(self):
        # Actualizo el valor
        self.m_current += 1
        # Devuelvo el valor antes de haberlo actualizado
        return self.m_current - 1

    def int(self):
        return self.m_current

def get_url_from_id(id):
    """
    Me espero el id en cadena, por si acaso hago la conversión
    """
    return 'https://www.filmaffinity.com/es/film' + str(id) + ".html"

def safe_get_url(url):
    #open with GET method
    resp = requests.get(url)
    # Caso 429: too many requests
    if resp.status_code == 429:
        return PassCaptcha(url)
    else: # No está contemplado el caso 404: not found
        return resp

def PassCaptcha(url):
    # abro un navegador para poder pasar el Captcha
    webbrowser.open(url)
    resp = requests.get(url)
    print("\nPor favor, entra en FilmAffinity y pasa el captcha por mí.")
    # Controlo que se haya pasado el Captcha
    while resp.status_code != 200:
        time.sleep(3) # intento recargar la página cada 3 segundos
        resp = requests.get(url)
    return resp

def set_cell_value(ws, line, col, value):
    cell = ws.cell(row = line, column=col)
    cell.value = value
    # Configuramos el estilo de la celda
    if (col == 5): # visionados. Ponemos punto de millar
        cell.number_format = '#,##0'
    elif (col == 9): # booleano mayor que
        cell.number_format = '0'
        cell.font = Font(name = 'SimSun', bold = True)
        cell.alignment=Alignment(horizontal='center', vertical='center')
    elif (col == 10):
        cell.number_format = '0.0'
    elif (col == 11 or col == 12): #reescala
        cell.number_format = '0.00'

def GetTotalFilms(resp):
    soup=BeautifulSoup(resp.text,'html.parser')
    # me espero que haya un único "value-box active-tab"
    mydivs = soup.find("a", {"class": "value-box active-tab"})
    stringNumber = mydivs.contents[3].contents[1]
    # Elimino el punto de los millares
    stringNumber = stringNumber.replace('.','')
    return int(stringNumber)

def IndexToLine(index, total):
    return total - index + 1

class Writer(object):
    def __init__(self, id, worksheet):
        # numero de usuario en string
        self.id_user = str(id)
        # Contador de peliculas
        self.film_index = 0
        # Numero de pagina actual
        self.page_index = 0

        # Descargo la propia página actual. Es una página "de fuera".
        self.soup_page = None
        # Lista de peliculas que hay en la página actual
        self.film_list = []
        #Inicializo las dos variables
        self.next_page()

        # Votaciones en total
        self.total_films = self.get_total_films()
        # Linea del excel en la que estoy escribiendo
        self.line = IndexLine(self.total_films)
        # Barra de progreso
        self.bar = None
        # Hoja de excel
        self.ws = worksheet

    def get_total_films(self):
        # me espero que haya un único "value-box active-tab"
        mydivs = self.soup_page.find("a", {"class": "value-box active-tab"})
        stringNumber = mydivs.contents[3].contents[1]
        # Elimino el punto de los millares
        stringNumber = stringNumber.replace('.','')
        return int(stringNumber)

    def get_list_url(self, page_index):
        sz_ans = 'https://www.filmaffinity.com/es/userratings.php?user_id=' + self.id_user + '&p='
        sz_ans = sz_ans + str(page_index) + '&orderby=4'

        return sz_ans

    def next_page(self):
        # Anavanzo a la siguiente página
        self.page_index += 1
        url = self.get_list_url(self.page_index)
        resp = safe_get_url(url)
        # Guardo la página ya parseada
        self.soup_page = BeautifulSoup(resp.text,'html.parser')
        # Leo todas las películas que haya en ella
        self.film_list = self.soup_page.findAll("div", {"class": "user-ratings-movie"})

    def read_watched(self):
        # Creo una barra de proceso
        self.bar = ProgressBar()

        # Itero hasta que haya leído todas las películas
        while (self.film_index < self.total_films):
            # Itero las películas en mi página actual
            for film in self.film_list:
                # Convierto lo leído de la página a un objeto película
                film = Pelicula(film)

                # Compruebo que su título sea válido
                if film.valid():
                    # Escribo sus datos en el excel
                    self.write_in_excel(film)

                # Paso a la siguiente película
                self.next_film()

            self.next_page()

    def write_in_excel(self, film):
        # Convierto el iterador en un entero
        line = self.line.int()
        # La votacion del usuario la leo desde fuera
        # no puedo leer la nota del usuario dentro de la ficha
        UserNote = film.user_note
        set_cell_value(self.ws, line, 2, int(UserNote))
        set_cell_value(self.ws, line, 10, str("=B" + str(line) + "+RAND()-0.5"))
        set_cell_value(self.ws, line, 11, "=(B" + str(line) + "-1)*10/9")
        # En la primera columna guardo la id para poder reconocerla
        n_id = film.id
        set_cell_value(self.ws, line, 1, int(n_id))
        film.get_time_and_FA()
        if (film.duracion != 0):
            # dejo la casilla en blanco si no logra leer ninguna duración de FA
            set_cell_value(self.ws, line, 4, film.duracion)
        if (film.nota_FA != 0):
            # dejo la casilla en blanco si no logra leer ninguna nota de FA
            set_cell_value(self.ws, line, 3, film.nota_FA)
            set_cell_value(self.ws, line, 6, "=ROUND(C" + str(line) + "*2, 0)/2")
            set_cell_value(self.ws, line, 7, "=B" + str(line) + "-C" + str(line))
            set_cell_value(self.ws, line, 8, "=ABS(G" + str(line) + ")")
            set_cell_value(self.ws, line, 9, "=IF($G" + str(line) + ">0,1,0.1)")
            set_cell_value(self.ws, line, 12, "=(C" + str(line) + "-1)*10/9")
        if (film.votantes_FA != 0):
            # dejo la casilla en blanco si no logra leer ninguna votantes
            set_cell_value(self.ws, line, 5, film.votantes_FA)

        # Como he escrito en el excel, paso a la línea siguiente
        self.line.get_current_line()

    def next_film(self):
        self.film_index += 1
        self.bar.update(self.film_index/self.total_films)

def get_user():
    Ids = {'Sasha': 1230513, 'Jorge': 1742789, 'Guillermo': 4627260, 'Daniel Gallego': 983049, 'Luminador': 7183467,
    'Will_llermo': 565861, 'Roger Peris': 3922745, 'Javi': 247783, 'El Feo': 867335, 'coleto': 527264}
    usuario = 'Jorge'
    print("Se van a importar los datos de ", usuario)
    inp = input("Espero Enter...")
    if inp and inp in Ids.keys():
        usuario = inp
        print("Se van a importar los datos de ", usuario)

    return usuario, Ids[usuario]


if __name__ == "__main__":
    usuario, id = get_user()
    Plantilla = 'Plantilla.xlsx'
    ExcelName = 'Sintaxis - ' + usuario + '.xlsx'
    workbook = load_workbook(Plantilla)
    worksheet = workbook[workbook.sheetnames[0]]
    writer = Writer(id, worksheet)
    writer.read_watched()
    workbook.save(ExcelName)
    workbook.close()
