import requests
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
# from openpyxl.styles import PatternFill, Border, Side, numbers
import sys
from bs4 import BeautifulSoup
import time
import datetime
import random
 
class Timer(object):
    def __init__(self):
        self.start = datetime.datetime.now()
 
    def remains(self, done):
        now  = datetime.datetime.now()
        left = (1 - done) * (now - self.start) / done
        sec = int(left.total_seconds())
        if sec < 60:
           return "{} seconds".format(sec)
        else:
           return "{} minutes".format(int(sec / 60))

def GetTimeAndFA(soup):

    # Comienzo a leer datos de la página
    l = soup.find(id="movie-rat-avg")
    try:
        # guardo la nota de FA en una ficha normal
        dNotaFA = float(l.attrs['content'])
    except:
        # caso en el que no hay nota de FA
        dNotaFA = 0
        return 0, 0, 0 # No pierdo el timepo leyendo otros datos
    
    l=soup.find(itemprop="ratingCount")
    try:
        # guardo la cantidad de votantes en una ficha normal
        nVotantes = l.attrs['content']
        # Elimino el punto de los millares
        nVotantes = nVotantes.replace('.','')
        nVotantes = int(nVotantes)
    except:
        # caso en el que no hay suficientes votantes
        nVotantes = 0
    
    l = soup.find(id = "left-column")
    try:
        duracion = l.find(itemprop="duration").contents[0]
    except:
        # caso en el que no está escrita la duración
        duracion = "0"
    # quito el sufijo min.
    duracion = int(duracion.split(' ', 1)[0])

    return dNotaFA, duracion, nVotantes

def SafeGetUrl():
    CodigoPelicula = random.randint(100000, 1000000)
    url = 'https://www.filmaffinity.com/es/film' + str(CodigoPelicula) + '.html'
    #open with GET method 
    resp = requests.get(url)
    # Caso 429: too many requests
    if resp.status_code == 429:
        resp = PassCaptcha(url)
    if resp.status_code == 404:
        return SafeGetUrl()
    else:
        return resp

def PassCaptcha(url):
    # abro un navegador para poder pasar el Captcha
    webbrowser.open(url)
    resp = requests.get(url)
    print("\nPor favor, entra en FilmAffinity y pasa el captcha por mí.")
    # Controlo que se haya pasado el Captcha
    while resp.status_code == 429:
        time.sleep(3) # intento recargar la página cada 3 segundos
        resp = requests.get(url)
    return resp
    
def SetCellValue(ws, line, col, value):
    cell = ws.cell(row = line, column=col)
    cell.value = value
    # Configuramos el estilo de la celda
    if (col == 5): # visionados. Ponemos punto de millar
        cell.number_format = '#,##0'
    elif (col == 9): # booleano mayor que
        cell.number_format = '0'
        cell.font = Font(name = 'SimSun', bold = True)
        cell.alignment=Alignment(horizontal='center', vertical='center')
    elif (col == 10):
        cell.number_format = '0.0'
    elif (col == 11 or col == 12): #reescala
        cell.number_format = '0.00'

def GetTotalFilms(resp):
    soup=BeautifulSoup(resp.text,'html.parser')
    # me espero que haya un único "value-box active-tab"
    mydivs = soup.find("a", {"class": "value-box active-tab"})
    stringNumber = mydivs.contents[3].contents[1]
    # Elimino el punto de los millares
    stringNumber = stringNumber.replace('.','')
    return int(stringNumber)

def update_progress(progress, timer):
    barLength = 20 # Modify this to change the length of the progress bar
    if isinstance(progress, int):
        progress = float(progress)
    block = int(round(barLength * progress))
    text = "\r[{0}] {1:.2f}% {2}".format( "="*block + " "*(barLength-block), progress*100, timer.remains(progress))
    sys.stdout.write(text)
    sys.stdout.flush()

def IndexToLine(index, total):
    return total - index + 1


def ReadWatched(ws):
    DataIndex = 0 # contador de peliculas
    notaFA = 0.0
    duracion = 0
    
    totalFilms = 500
    line = IndexToLine(DataIndex, totalFilms) # linea de excel en la que estoy escribiendo
    timer = Timer()
    while (DataIndex < totalFilms):
        resp = SafeGetUrl()
        # we need a parser,Python built-in HTML parser is enough . 
        soup = BeautifulSoup(resp.text,'html.parser')
        
        #Entro a la ficha y leo votacion popular, duracion y votantes
        notaFA, duracion, votantes = GetTimeAndFA(soup)
        if notaFA == 0:
            continue
        if (duracion != 0):
            # dejo la casilla en blanco si no logra leer ninguna duración de FA
            SetCellValue(ws, line, 4, duracion)
        if (notaFA != 0):
            # dejo la casilla en blanco si no logra leer ninguna nota de FA
            SetCellValue(ws, line, 3, notaFA)
            SetCellValue(ws, line, 6, "=ROUND(C" + str(line) + "*2, 0)/2")
            SetCellValue(ws, line, 7, "=B" + str(line) + "-C" + str(line))
            SetCellValue(ws, line, 8, "=ABS(G" + str(line) + ")")
            SetCellValue(ws, line, 9, "=IF($G" + str(line) + ">0,1,0.1)")
            SetCellValue(ws, line, 12, "=(C" + str(line) + "-1)*10/9")
        if (votantes != 0):
            # dejo la casilla en blanco si no logra leer ninguna votantes
            SetCellValue(ws, line, 5, votantes)
        DataIndex +=1
        # actualizo la barra de progreso
        update_progress(DataIndex/totalFilms, timer)
        # actualizo la linea de escritura en excel
        line = IndexToLine(DataIndex, totalFilms)



if __name__ == "__main__":
    Ids = {'Sasha': 1230513, 'Jorge': 1742789, 'Guillermo': 4627260, 'Daniel Gallego': 983049, 'Luminador': 7183467,
    'Will_llermo': 565861, 'Roger Peris': 3922745, 'Javi': 247783}
    usuario = 'Sasha'
    print("Se van a importar los datos de ", usuario)
    input("Espero Enter...")
    Plantilla = 'Plantilla.xlsx'
    ExcelName = 'Random.xlsx'
    workbook = load_workbook(Plantilla)
    worksheet = workbook[workbook.sheetnames[0]]
    ReadWatched(worksheet)
    workbook.save(ExcelName)
    workbook.close()
