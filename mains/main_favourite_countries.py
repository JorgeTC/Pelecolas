from collections import Counter

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import __init__
from src.aux_res_directory import get_res_folder
from src.safe_url import safe_get_url


def make_url(year: int, min_votes='100', min_duration='0', max_duration='+120') -> str:
    url = 'https://www.filmaffinity.com/es/topgen.php?'
    # Generos
    url += 'genres=&'
    url += 'chv=0&'
    # Sorted
    url += 'orderby=avg&'
    # Type filter
    url += 'movietype=movie%7C&'
    # Country filter
    url += 'country=&'
    # Only one year
    url += f'fromyear={year}&toyear={year}&'
    # Min number of votes
    number_of_votes = {'20': 0, '100': 1, '500': 2, '1000': 3}
    url += f'ratingcount={number_of_votes[min_votes]}&'

    duration = {'0': 0, '60': 1, '90': 2, '120': 3, '+120': 4}
    # Min duration
    url += f'runtimemin={duration[min_duration]}&'
    # Max duration
    url += f'runtimemax={duration[max_duration]}'

    return url


COUNTRIES_PER_YEAR = 15


def extract_flags(content: str) -> dict[str, int]:

    # Trim until the section we want
    position = content.find('id="top-movies"')
    content = content[position:]

    countries = []
    for _ in range(15):
        # Find the next country
        if ((position := content.find('<img class="nflag"')) == -1):
            break
        # Trim string
        content = content[position:]
        # Find the label for the country
        position = content.find('alt="')
        # Trim the label
        content = content[position + len('alt="'):]

        # Extract the country
        position = content.find('"')
        countries.append(content[:position])

        # Trim string
        content = content[position+1:]

    return dict(Counter(countries))


def main():

    flags: dict[int, dict[str, int]] = {}

    # Importo los módulos del programa cuando la configuración ya está settada
    for year in range(1915, 2023):
        print(f"Reading year {year}")
        url = make_url(year)
        response = safe_get_url(url)

        flags_current_year = extract_flags(response.text)

        flags[year] = flags_current_year

    # Me guardo todos los países
    to_write: dict[str, list[int]] = {}
    for year in flags:
        for country in flags[year]:
            to_write[country]: list[int] = []

    for country in to_write:
        for year in flags:
            to_write[country].append(flags[year].get(country, 0))

    excel_path = get_res_folder('countries.xlsx')
    wb = load_workbook(excel_path)
    # Abro la primera de las hojas, es la única en la que escribo
    ws: Worksheet = wb[wb.sheetnames[0]]

    # Escribo primero todos los paises
    for col, country in enumerate(to_write):
        cell = ws.cell(row=1, column=col+2)
        cell.value = country

    # Escribo todos los años
    for row, year in enumerate(flags):
        cell = ws.cell(row=row+2, column=1)
        cell.value = year

    # Escribo el relleno de la tabla
    for col, country in enumerate(to_write):
        for row, number in enumerate(to_write[country]):
            cell = ws.cell(row=row+2, column=col+2)
            cell.value = number

    wb.save(excel_path)
    wb.close()

    print("End")


if __name__ == "__main__":
    main()
