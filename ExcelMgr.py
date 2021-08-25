
from pathlib import Path
from openpyxl import load_workbook

class ExcelMgr(object):
    def __init__(self, usuario):
        # El nombre de la plantilla siempre es el mismo
        Plantilla = 'Plantilla.xlsx'

        sz_curr_folder = Path(__file__).resolve().parent
        Plantilla = sz_curr_folder / Plantilla

        self.wb = load_workbook(Plantilla)
        self.ws = self.wb[self.wb.sheetnames[0]]

        self.ExcelName = 'Sintaxis - ' + usuario + '.xlsx'

    def get_worksheet(self):
        return self.ws

    def save_wb(self):

        sz_curr_folder = Path(__file__).resolve().parent
        # Me voy hasta la carpeta Reseñas
        sz_curr_folder = sz_curr_folder.parent.parent
        # Entro a la carpeta películas
        sz_curr_folder = sz_curr_folder / "Películas"

        self.wb.save(sz_curr_folder / self.ExcelName)
        self.wb.close()