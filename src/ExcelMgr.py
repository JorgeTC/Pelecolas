
from pathlib import Path
from openpyxl import load_workbook

class ExcelMgr(object):

    SZ_FILE_NAME = 'Sintaxis - {}.xlsx'
    SZ_TEMPLATE_NAME = 'Plantilla.xlsx'

    def __init__(self, usuario):
        # Abro la carpeta donde está la plantilla
        # Carpeta source
        sz_curr_folder = Path(__file__).parent
        # Carpeta del proyecto
        sz_curr_folder = sz_curr_folder.parent
        # Carpeta de recursos
        sz_curr_folder = sz_curr_folder / "res"
        # Readdata
        sz_curr_folder = sz_curr_folder / "Readdata"
        # Construyo el path completo del archivo
        Plantilla = sz_curr_folder / self.SZ_TEMPLATE_NAME

        # Abro el archivo excel
        self.wb = load_workbook(Plantilla)
        # Abro la primera de las hojas, es la única en la que escribo
        self.ws = self.wb[self.wb.sheetnames[0]]

        # Contruyo el nombre con el que voy a guardar el excel
        self.ExcelName = self.SZ_FILE_NAME.format(usuario)

    def get_worksheet(self):
        return self.ws

    def save_wb(self, path):
        # Me han pasado la carpeta de destino como argumento
        self.wb.save(path / self.ExcelName)
        # Cierro el archivo excel
        self.wb.close()
