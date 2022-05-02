from openpyxl import load_workbook

from src.aux_res_directory import get_res_folder
from src.dlg_config import Config, Section, Param


class ExcelMgr():

    SZ_FILE_NAME = 'Sintaxis - {}.xlsx'
    SZ_TEMPLATE_NAME = 'Plantilla.xlsx'

    def __init__(self, usuario):
        # Abro la carpeta donde está la plantilla
        Plantilla = get_res_folder("Readdata", self.SZ_TEMPLATE_NAME)

        # Abro el archivo excel
        self.wb = load_workbook(Plantilla)
        # Abro la primera de las hojas, es la única en la que escribo
        self.ws = self.wb[self.wb.sheetnames[0]]

        # Construyo el nombre con el que voy a guardar el excel
        self.ExcelName = self.SZ_FILE_NAME.format(usuario)

        # Cargo la carpeta donde se guardará
        self.output_path = Config.get_folder_path(
            Section.READDATA, Param.OUTPUT_EXCEL)

    def get_worksheet(self):
        return self.ws

    def save_wb(self):
        # Me han pasado la carpeta de destino como argumento
        self.wb.save(self.output_path / self.ExcelName)
        # Cierro el archivo excel
        self.wb.close()
